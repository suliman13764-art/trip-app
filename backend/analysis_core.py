from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import pdfplumber
from openpyxl import load_workbook


GPS_DATETIME_FORMATS = [
    "%d-%m-%Y %H:%M:%S",
    "%d-%m-%Y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
]
WEBTRACK_DT_RE = re.compile(r"^(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}(?::\d{2})?)(.*)$")
RUN_RE = re.compile(r"\b\d{4}/\d{4}\b")
STOP_ACTION_RE = re.compile(r"(?:Total\s+\d+\s+stop\s+)?(\d{2})\s+(AFLEV|BAG)\s+(\d{2}:\d{2})")
STOP_NOTICE_RE = re.compile(r"\b(\d{2})\(\+\d{8,}\)")
PRIMARY_ORDER_RE = re.compile(r"\b(\d{4})\s+\d{6}\s+PLA\b")
STOP_ORDER_RE = re.compile(r"(?:AFLEV|BAG)\s+\d{2}:\d{2}\s+(\d{4,})\b")
AGE_ADDRESS_RE = re.compile(r"(?:[MK]\d{1,3}ÅR)\s+(.+?)\s+PN:(\d{4})")
POSTAL_ADDRESS_RE = re.compile(
    r"PN:(\d{4})\s+(.+?)(?=\s+(?:OBS|INTET|Tlf\.?|tlf\.?|\(\d{2}:\d{2}\)|\d+\s+PER|FLERE\s+STOP|$))"
)
TIME_IN_TEXT_RE = re.compile(r"\b\d{1,2}:\d{2}\b")
SUMMARY_DRIVE_RE = re.compile(r"Køretid by:\s*(\d+)\s+Land:\s*(\d+)")
SUMMARY_WAIT_RE = re.compile(r"Ventetid by:\s*(\d+)\s+Land:\s*(\d+)")
WORK_MESSAGE_TYPES = {"route_end", "drive_home", "bag", "aflev", "aflev/bag", "general"}
STOP_ADDRESS_SIMILARITY_MIN = 0.25


@dataclass
class GPSPoint:
    timestamp: datetime
    latitude: float
    longitude: float
    speed: float | None
    status: str
    address: str
    distance_to_home_m: float = 0.0
    inside_home_zone: bool = False


@dataclass
class Segment:
    start_time: datetime
    end_time: datetime | None
    start_point: dict[str, Any]
    end_point: dict[str, Any] | None
    closure_reason: str
    is_estimated_end: bool = False
    stops: list[int] | None = None
    private_trip_split: bool = False


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    text = text.replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None



def parse_datetime_value(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    for fmt in GPS_DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)



def combine_date_and_clock(source_date: datetime, clock_text: str | None) -> datetime | None:
    if not source_date or not clock_text:
        return None
    try:
        parsed_clock = datetime.strptime(clock_text, "%H:%M")
    except ValueError:
        return None
    return source_date.replace(hour=parsed_clock.hour, minute=parsed_clock.minute, second=0, microsecond=0)



def haversine_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6_371_000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(delta_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    )
    return 2 * radius * math.asin(math.sqrt(a))



def _detect_csv_separator(sample: str) -> str:
    candidates = [";", ",", "\t", "|"]
    counts = {sep: sample.count(sep) for sep in candidates}
    return max(counts, key=counts.get)



def parse_gps_file(file_path: str | Path) -> list[GPSPoint]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xls"}:
        raise ValueError(f"Unsupported GPS file type: {suffix}")

    if suffix == ".csv":
        encodings = ["utf-8-sig", "utf-8", "latin1", "cp1252"]
        last_error = None
        for encoding in encodings:
            try:
                with open(path, "r", encoding=encoding) as handle:
                    sample = handle.read(5000)
                sep = _detect_csv_separator(sample)
                frame = pd.read_csv(path, sep=sep, encoding=encoding, engine="python")
                break
            except Exception as exc:  # noqa: BLE001
                last_error = exc
        else:
            raise ValueError(f"Unable to read GPS CSV file: {last_error}")
    else:
        frame = pd.read_excel(path)

    frame = frame.dropna(axis=1, how="all")
    if frame.empty:
        raise ValueError("GPS file contains no usable rows")

    original_columns = list(frame.columns)
    normalized_columns = {str(col).strip().lower(): col for col in original_columns}
    datetime_col = next(
        (
            normalized_columns[name]
            for name in normalized_columns
            if any(token in name for token in ["date", "time", "timestamp", "dato"])
        ),
        None,
    )
    lat_col = next(
        (
            normalized_columns[name]
            for name in normalized_columns
            if any(token in name for token in ["latitude", "lat"])
        ),
        None,
    )
    lon_col = next(
        (
            normalized_columns[name]
            for name in normalized_columns
            if any(token in name for token in ["longitude", "long", "lon"])
        ),
        None,
    )
    speed_col = next((normalized_columns[name] for name in normalized_columns if "speed" in name), None)
    status_col = next((normalized_columns[name] for name in normalized_columns if "status" in name), None)
    address_col = next((normalized_columns[name] for name in normalized_columns if "address" in name), None)

    if not datetime_col or not lat_col or not lon_col:
        raise ValueError(
            "GPS file must contain timestamp/date, latitude and longitude columns. "
            f"Detected columns: {original_columns}"
        )

    points: list[GPSPoint] = []
    for _, row in frame.iterrows():
        timestamp = parse_datetime_value(row.get(datetime_col))
        latitude = _safe_float(row.get(lat_col))
        longitude = _safe_float(row.get(lon_col))
        if timestamp is None or latitude is None or longitude is None:
            continue
        speed = _safe_float(row.get(speed_col)) if speed_col else None
        points.append(
            GPSPoint(
                timestamp=timestamp,
                latitude=latitude,
                longitude=longitude,
                speed=speed,
                status=str(row.get(status_col) or "").strip(),
                address=str(row.get(address_col) or "").strip(),
            )
        )

    if not points:
        raise ValueError("GPS file contained rows, but no valid timestamp/coordinate records could be parsed")

    points.sort(key=lambda point: point.timestamp)
    return points



def _extract_lines_from_pdf(path: Path) -> list[str]:
    lines: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(text.splitlines())
    return lines



def _extract_lines_from_excel(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" ".join(values))
    return lines



def _clean_webtrack_lines(lines: Iterable[str]) -> list[str]:
    cleaned: list[str] = []
    for raw_line in lines:
        line = re.sub(r"\s+", " ", str(raw_line).strip())
        if not line:
            continue
        if line.startswith("https://") or re.fullmatch(r"\d+/\d+", line):
            continue
        if re.match(r"^\d{1,2}/\d{1,2}/\d{2},", line):
            continue
        cleaned.append(line)
    return cleaned



def parse_webtrack_file(file_path: str | Path) -> list[dict[str, Any]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        lines = _extract_lines_from_pdf(path)
    elif suffix in {".xlsx", ".xls", ".csv"}:
        if suffix == ".csv":
            frame = pd.read_csv(path, header=None, encoding="latin1")
            lines = [" ".join(str(cell).strip() for cell in row if pd.notna(cell) and str(cell).strip()) for row in frame.values]
        else:
            lines = _extract_lines_from_excel(path)
    else:
        raise ValueError(f"Unsupported WebTrack file type: {suffix}")

    cleaned_lines = _clean_webtrack_lines(lines)
    records: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    def push_current() -> None:
        if current is None:
            return
        payload_lines = [part for part in current["parts"] if part]
        raw_text = " ".join(payload_lines).strip()
        run_numbers = sorted(set(RUN_RE.findall(raw_text)))
        stop_matches = [
            {
                "stop_number": match[0],
                "action": match[1],
                "planned_time": match[2],
                "completion_datetime": combine_date_and_clock(current["timestamp"], match[2]),
            }
            for match in STOP_ACTION_RE.findall(raw_text)
        ]
        notice_stops = sorted(set(STOP_NOTICE_RE.findall(raw_text)))

        primary_order = PRIMARY_ORDER_RE.search(raw_text)
        stop_order = STOP_ORDER_RE.search(raw_text)
        order_number = None
        if primary_order:
            order_number = primary_order.group(1)
        elif stop_order:
            order_number = stop_order.group(1)
        else:
            loose_numbers = [
                number
                for number in re.findall(r"\b\d{4}\b", raw_text)
                if number not in {part for run in run_numbers for part in run.split("/")}
            ]
            order_number = loose_numbers[0] if loose_numbers else None

        v_lobe_slut = re.search(r"V\.LØBE SLUT\s+(\d{1,2}:\d{2})", raw_text)
        homezone_message = re.search(r"KØRE TIL HJEMZONE", raw_text)

        extracted_address = None
        post_match = POSTAL_ADDRESS_RE.search(raw_text)
        age_match = AGE_ADDRESS_RE.search(raw_text)
        if post_match:
            postal_code, address_part = post_match.groups()
            extracted_address = f"{address_part.strip()}, {postal_code}".strip(", ")
        elif age_match:
            address_part, postal_code = age_match.groups()
            extracted_address = f"{address_part.strip()}, {postal_code}".strip(", ")

        message_type = "general"
        if "V.LØBE SLUT" in raw_text:
            message_type = "route_end"
        elif "KØRE TIL HJEMZONE" in raw_text:
            message_type = "drive_home"
        elif stop_matches:
            actions = sorted({match["action"] for match in stop_matches})
            message_type = "/".join(actions).lower()
        elif "NewOrder" in raw_text:
            message_type = "new_order"

        completion_times = [
            match["completion_datetime"]
            for match in stop_matches
            if match["action"] == "AFLEV" and match["completion_datetime"] is not None
        ]
        drive_summary_match = SUMMARY_DRIVE_RE.search(raw_text)
        wait_summary_match = SUMMARY_WAIT_RE.search(raw_text)
        final_summary = None
        if drive_summary_match or wait_summary_match:
            final_summary = {
                "drive_city_minutes": int(drive_summary_match.group(1)) if drive_summary_match else 0,
                "drive_land_minutes": int(drive_summary_match.group(2)) if drive_summary_match else 0,
                "wait_city_minutes": int(wait_summary_match.group(1)) if wait_summary_match else 0,
                "wait_land_minutes": int(wait_summary_match.group(2)) if wait_summary_match else 0,
            }
            final_summary["total_drive_minutes"] = final_summary["drive_city_minutes"] + final_summary["drive_land_minutes"]
            final_summary["total_wait_minutes"] = final_summary["wait_city_minutes"] + final_summary["wait_land_minutes"]
            final_summary["afregnet_minutes"] = final_summary["total_drive_minutes"] + final_summary["total_wait_minutes"]

        stop_events = [
            {
                "stop_number": int(match["stop_number"]) if str(match["stop_number"]).isdigit() else match["stop_number"],
                "action": match["action"],
                "planned_time": match["planned_time"],
                "planned_datetime": match["completion_datetime"],
                "record_timestamp": current["timestamp"],
                "address": extracted_address,
                "order_number": order_number,
                "raw_text": raw_text,
            }
            for match in stop_matches
        ]

        records.append(
            {
                "timestamp": current["timestamp"],
                "raw_text": raw_text,
                "run_numbers": run_numbers,
                "stop_matches": stop_matches,
                "stop_events": stop_events,
                "notice_stops": notice_stops,
                "order_number": order_number,
                "message_type": message_type,
                "v_lobe_slut_time": v_lobe_slut.group(1) if v_lobe_slut else None,
                "contains_drive_home": bool(homezone_message),
                "extracted_address": extracted_address,
                "service_times": sorted(set(TIME_IN_TEXT_RE.findall(raw_text))),
                "completion_times": completion_times,
                "final_summary": final_summary,
                "is_meaningful": message_type != "new_order",
            }
        )

    for line in cleaned_lines:
        match = WEBTRACK_DT_RE.match(line)
        if match:
            push_current()
            timestamp = parse_datetime_value(match.group(1))
            current = {"timestamp": timestamp, "parts": [match.group(2).strip()] if match.group(2).strip() else []}
        elif current is not None:
            current["parts"].append(line)
    push_current()

    records = [record for record in records if record["timestamp"] is not None]
    records.sort(key=lambda record: record["timestamp"])
    if not records:
        raise ValueError("No timestamped WebTrack records could be extracted from the file")
    return records



def summarize_webtrack(records: list[dict[str, Any]]) -> dict[str, Any]:
    run_numbers = sorted({run for record in records for run in record["run_numbers"]})
    stop_numbers = sorted(
        {
            int(match["stop_number"])
            for record in records
            for match in record["stop_matches"]
            if str(match["stop_number"]).isdigit()
        }
        | {int(stop) for record in records for stop in record["notice_stops"] if str(stop).isdigit()}
    )
    primary_order_numbers = [record["order_number"] for record in records if record["order_number"]]
    primary_order = max(set(primary_order_numbers), key=primary_order_numbers.count) if primary_order_numbers else None

    last_meaningful_event = next(
        (
            record
            for record in reversed(records)
            if record["message_type"] in {"route_end", "drive_home", "bag", "aflev", "aflev/bag"}
        ),
        records[-1],
    )
    last_order_with_address = next(
        (
            record
            for record in reversed(records)
            if record["extracted_address"] and record["message_type"] in {"bag", "aflev", "aflev/bag"}
        ),
        None,
    )
    completion_events = [
        {
            "timestamp": completion_time,
            "stop_number": match["stop_number"],
            "order_number": record["order_number"],
            "raw_text": record["raw_text"],
        }
        for record in records
        for match in record["stop_matches"]
        for completion_time in [match["completion_datetime"]]
        if match["action"] == "AFLEV" and completion_time is not None
    ]
    completion_events.sort(key=lambda item: item["timestamp"])
    meaningful_event_times = [
        {
            "timestamp": record["timestamp"],
            "message_type": record["message_type"],
            "raw_text": record["raw_text"],
        }
        for record in records
        if record["message_type"] in WORK_MESSAGE_TYPES
    ]
    stop_events = [event for record in records for event in record.get("stop_events", []) if event.get("planned_datetime")]
    stop_events.sort(key=lambda event: event["planned_datetime"])
    final_summary_record = next((record for record in reversed(records) if record.get("final_summary")), None)
    final_summary = final_summary_record.get("final_summary") if final_summary_record else None

    return {
        "run_numbers": run_numbers,
        "primary_run_number": run_numbers[0] if run_numbers else None,
        "primary_order_number": primary_order,
        "stop_numbers": stop_numbers,
        "stop_count": len(stop_numbers),
        "first_record_time": records[0]["timestamp"],
        "last_record_time": records[-1]["timestamp"],
        "last_meaningful_event": last_meaningful_event,
        "last_order_with_address": last_order_with_address,
        "v_lobe_slut_times": [record["v_lobe_slut_time"] for record in records if record["v_lobe_slut_time"]],
        "drive_home_messages": [record for record in records if record["contains_drive_home"]],
        "completion_events": completion_events,
        "meaningful_event_times": meaningful_event_times,
        "stop_events": stop_events,
        "final_summary": final_summary,
        "final_summary_record_time": final_summary_record["timestamp"] if final_summary_record else None,
    }


def _normalize_address_tokens(value: str | None) -> set[str]:
    if not value:
        return set()
    normalized = value.lower().replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    tokens = set(re.findall(r"[a-z0-9]+", normalized))
    ignored = {"danmark", "pn", "tv", "st", "th", "sal", "vej", "gade"}
    return {token for token in tokens if token not in ignored and len(token) > 1}


def _address_similarity(left: str | None, right: str | None) -> float:
    left_tokens = _normalize_address_tokens(left)
    right_tokens = _normalize_address_tokens(right)
    if not left_tokens or not right_tokens:
        return 0.0
    intersection = left_tokens & right_tokens
    union = left_tokens | right_tokens
    return len(intersection) / max(1, len(union))


def infer_actual_stop_timing(stop_event: dict[str, Any], gps_points: list[GPSPoint]) -> dict[str, Any]:
    planned_datetime = stop_event.get("planned_datetime")
    if planned_datetime is None:
        return {
            "actual_datetime": None,
            "delay_minutes": None,
            "confidence": "low",
            "reason_basis": "Missing planned stop time",
            "matched_gps_point": None,
            "stationary_minutes": 0,
        }

    candidates: list[tuple[float, float, GPSPoint]] = []
    for point in gps_points:
        time_diff = abs((point.timestamp - planned_datetime).total_seconds()) / 60
        if time_diff > 240:
            continue
        similarity = _address_similarity(stop_event.get("address"), point.address)
        if similarity < STOP_ADDRESS_SIMILARITY_MIN and time_diff > 45:
            continue
        speed_penalty = 0 if (point.speed is not None and point.speed <= 10) or point.status.lower() in {"stop", "paused", "parked"} else 6
        score = similarity * 100 - time_diff - speed_penalty
        candidates.append((score, time_diff, point))

    matched_point = max(candidates, key=lambda item: item[0])[2] if candidates else min(
        gps_points,
        key=lambda point: abs((point.timestamp - planned_datetime).total_seconds()),
    )
    similarity = _address_similarity(stop_event.get("address"), matched_point.address)
    neighborhood = [
        point
        for point in gps_points
        if abs((point.timestamp - matched_point.timestamp).total_seconds()) <= 15 * 60
        and ((point.speed is not None and point.speed <= 5) or point.status.lower() in {"stop", "paused", "parked"})
    ]
    stationary_minutes = 0
    if neighborhood:
        stationary_minutes = max(0, round((neighborhood[-1].timestamp - neighborhood[0].timestamp).total_seconds() / 60))

    delay_minutes = round((matched_point.timestamp - planned_datetime).total_seconds() / 60)
    confidence = "high" if similarity >= 0.45 else "medium" if similarity >= 0.25 else "low"
    reason_basis = f"GPS address match {similarity:.2f}; stationary {stationary_minutes} min"
    return {
        "actual_datetime": matched_point.timestamp,
        "delay_minutes": delay_minutes,
        "confidence": confidence,
        "reason_basis": reason_basis,
        "matched_gps_point": _point_to_dict(matched_point),
        "stationary_minutes": stationary_minutes,
    }


def analyze_delay_points(webtrack_summary: dict[str, Any], gps_points: list[GPSPoint]) -> dict[str, Any]:
    stop_delay_analysis: list[dict[str, Any]] = []
    previous_actual = None
    previous_planned = None

    for stop_event in webtrack_summary.get("stop_events", []):
        inferred = infer_actual_stop_timing(stop_event, gps_points)
        actual_datetime = inferred["actual_datetime"]
        planned_datetime = stop_event.get("planned_datetime")
        delay_minutes = inferred["delay_minutes"]
        reason = "No major delay detected"
        if delay_minutes is None:
            reason = "Insufficient data"
        elif delay_minutes > 5:
            if inferred["stationary_minutes"] >= 5:
                reason = "waiting time at stop"
            elif previous_actual and previous_planned:
                planned_gap = (planned_datetime - previous_planned).total_seconds() / 60
                actual_gap = (actual_datetime - previous_actual).total_seconds() / 60 if actual_datetime else planned_gap
                reason = "sequence gap between stops" if actual_gap - planned_gap > 10 else "traffic / travel time"
            else:
                reason = "traffic / travel time"
        elif delay_minutes < -5:
            reason = "ahead of planned stop time"

        record = {
            "stop_number": stop_event.get("stop_number"),
            "stop_type": stop_event.get("action"),
            "planned_time": planned_datetime.isoformat(sep=" ") if planned_datetime else None,
            "actual_time": actual_datetime.isoformat(sep=" ") if actual_datetime else None,
            "delay_minutes": delay_minutes,
            "delay_only_minutes": max(0, delay_minutes) if delay_minutes is not None else None,
            "reason": reason,
            "confidence": inferred["confidence"],
            "address": stop_event.get("address"),
            "reason_basis": inferred["reason_basis"],
            "matched_gps_point": inferred["matched_gps_point"],
        }
        stop_delay_analysis.append(record)
        previous_actual = actual_datetime or previous_actual
        previous_planned = planned_datetime or previous_planned

    significant = [item for item in stop_delay_analysis if (item.get("delay_only_minutes") or 0) > 5]
    main_delay = max(significant, key=lambda item: item["delay_only_minutes"]) if significant else None
    return {
        "stop_delay_analysis": stop_delay_analysis,
        "main_delay": main_delay,
        "significant_delays": significant,
    }


def derive_home_center_for_poc(points: list[GPSPoint]) -> tuple[float, float, str]:
    tail = points[-12:] if len(points) >= 12 else points
    stationary_tail = [
        point
        for point in tail
        if (point.speed is not None and point.speed <= 1)
        or point.status.lower() in {"stop", "paused", "parked"}
    ]
    candidates = stationary_tail or tail
    latitude = sum(point.latitude for point in candidates) / len(candidates)
    longitude = sum(point.longitude for point in candidates) / len(candidates)
    return latitude, longitude, "derived_from_final_stationary_cluster"



def _point_to_dict(point: GPSPoint) -> dict[str, Any]:
    return {
        "timestamp": point.timestamp.isoformat(sep=" "),
        "latitude": round(point.latitude, 6),
        "longitude": round(point.longitude, 6),
        "distance_to_home_m": round(point.distance_to_home_m, 1),
        "speed": point.speed,
        "status": point.status,
        "address": point.address,
        "inside_home_zone": point.inside_home_zone,
    }



def detect_segments(
    points: list[GPSPoint],
    home_latitude: float,
    home_longitude: float,
    radius_meters: float = 300,
    min_departure_points: int = 3,
    min_return_points: int = 3,
    min_return_dwell_minutes: int = 10,
) -> tuple[list[Segment], dict[str, Any]]:
    if not points:
        return [], {"closest_distance_m": None, "reason": "No GPS points"}

    for point in points:
        point.distance_to_home_m = haversine_meters(point.latitude, point.longitude, home_latitude, home_longitude)
        point.inside_home_zone = point.distance_to_home_m <= radius_meters

    runs: list[dict[str, Any]] = []
    current_points = [points[0]]
    current_inside = points[0].inside_home_zone
    for point in points[1:]:
        if point.inside_home_zone == current_inside:
            current_points.append(point)
            continue
        runs.append({"inside": current_inside, "points": current_points})
        current_points = [point]
        current_inside = point.inside_home_zone
    runs.append({"inside": current_inside, "points": current_points})

    segments: list[Segment] = []
    open_segment: Segment | None = None
    rejected_returns: list[dict[str, Any]] = []
    for index, run in enumerate(runs):
        run_points = run["points"]
        if run["inside"]:
            if open_segment is None:
                continue
            if len(run_points) < min_return_points:
                rejected_returns.append({
                    "candidate_entry_time": run_points[0].timestamp.isoformat(sep=" "),
                    "reason": "Too few consecutive inside points",
                })
                continue
            dwell_seconds = (run_points[-1].timestamp - run_points[0].timestamp).total_seconds()
            has_meaningful_stop = any(
                ((point.speed is not None and point.speed <= 1) or point.status.lower() in {"stop", "paused", "parked"})
                for point in run_points
            )
            reaches_file_end = index == len(runs) - 1
            valid_return = (
                dwell_seconds >= min_return_dwell_minutes * 60
                or has_meaningful_stop
                or (reaches_file_end and len(run_points) >= min_return_points and (has_meaningful_stop or dwell_seconds >= 180))
            )
            if valid_return:
                open_segment.end_time = run_points[0].timestamp
                open_segment.end_point = _point_to_dict(run_points[0])
                open_segment.closure_reason = (
                    "valid_home_return_with_dwell" if dwell_seconds >= min_return_dwell_minutes * 60 else "valid_home_return_with_stop_or_end_of_data"
                )
                segments.append(open_segment)
                open_segment = None
            else:
                rejected_returns.append({
                    "candidate_entry_time": run_points[0].timestamp.isoformat(sep=" "),
                    "reason": f"Inside run did not reach dwell threshold; dwell_seconds={int(dwell_seconds)}",
                })
            continue

        if len(run_points) < min_departure_points:
            continue
        if open_segment is None:
            open_segment = Segment(
                start_time=run_points[0].timestamp,
                end_time=None,
                start_point=_point_to_dict(run_points[0]),
                end_point=None,
                closure_reason="departure_confirmed_after_consecutive_outside_points",
                stops=[],
            )

    debug = {
        "closest_distance_m": round(min(point.distance_to_home_m for point in points), 1),
        "runs": [
            {
                "inside": run["inside"],
                "point_count": len(run["points"]),
                "start": run["points"][0].timestamp.isoformat(sep=" "),
                "end": run["points"][-1].timestamp.isoformat(sep=" "),
                "duration_seconds": int((run["points"][-1].timestamp - run["points"][0].timestamp).total_seconds()),
                "min_distance_m": round(min(point.distance_to_home_m for point in run["points"]), 1),
                "max_distance_m": round(max(point.distance_to_home_m for point in run["points"]), 1),
            }
            for run in runs
        ],
        "rejected_returns": rejected_returns,
    }
    return segments, debug



def normalize_private_trip_overrides(private_trip_overrides: list[dict[str, Any]] | None, analysis_date: datetime) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, trip in enumerate(private_trip_overrides or []):
        start = trip.get("start_time")
        end = trip.get("end_time")
        start_dt = parse_datetime_value(start) or combine_date_and_clock(analysis_date, str(start))
        end_dt = parse_datetime_value(end) or combine_date_and_clock(analysis_date, str(end))
        if not start_dt or not end_dt or end_dt <= start_dt:
            continue
        normalized.append(
            {
                "id": trip.get("id") or f"private-trip-{index + 1}",
                "start_time": start_dt,
                "end_time": end_dt,
                "source": trip.get("source", "manual"),
                "notes": trip.get("notes", ""),
                "confidence": trip.get("confidence", "confirmed"),
            }
        )
    normalized.sort(key=lambda item: item["start_time"])
    return normalized



def find_last_completion_before(webtrack_summary: dict[str, Any], boundary: datetime, segment_start: datetime) -> dict[str, Any] | None:
    candidates = [
        event for event in webtrack_summary["completion_events"] if segment_start <= event["timestamp"] <= boundary
    ]
    return candidates[-1] if candidates else None



def infer_resume_after(
    private_trip_end: datetime,
    gps_points: list[GPSPoint],
    webtrack_summary: dict[str, Any],
) -> tuple[datetime | None, dict[str, Any]]:
    next_gps = next(
        (
            point
            for point in gps_points
            if point.timestamp >= private_trip_end
            and ((point.speed is not None and point.speed > 1) or point.status.lower() in {"start", "moving"})
        ),
        None,
    )
    next_work_event = next(
        (
            event
            for event in webtrack_summary["meaningful_event_times"]
            if event["timestamp"] >= private_trip_end and event["message_type"] in WORK_MESSAGE_TYPES
        ),
        None,
    )

    resume_candidates = [candidate for candidate in [next_gps.timestamp if next_gps else None, next_work_event["timestamp"] if next_work_event else None] if candidate]
    resume_time = min(resume_candidates) if resume_candidates else None
    debug = {
        "next_gps_resume": next_gps.timestamp.isoformat(sep=" ") if next_gps else None,
        "next_work_event_resume": next_work_event["timestamp"].isoformat(sep=" ") if next_work_event else None,
        "selected_resume": resume_time.isoformat(sep=" ") if resume_time else None,
    }
    return resume_time, debug



def attach_segment_stop_ranges(segments: list[Segment], webtrack_summary: dict[str, Any]) -> None:
    completion_events = webtrack_summary["completion_events"]
    for segment in segments:
        if not segment.end_time:
            segment.stops = []
            continue
        segment.stops = sorted(
            {
                int(event["stop_number"])
                for event in completion_events
                if event["timestamp"] >= segment.start_time and event["timestamp"] <= segment.end_time and str(event["stop_number"]).isdigit()
            }
        )



def apply_private_trip_overlays(
    segments: list[Segment],
    webtrack_summary: dict[str, Any],
    gps_points: list[GPSPoint],
    private_trips: list[dict[str, Any]],
) -> tuple[list[Segment], list[dict[str, Any]], list[dict[str, Any]]]:
    if not private_trips:
        attach_segment_stop_ranges(segments, webtrack_summary)
        return segments, [], []

    adjusted_segments = segments[:]
    adjustments: list[dict[str, Any]] = []
    applied_private_trips: list[dict[str, Any]] = []

    for private_trip in private_trips:
        next_segments: list[Segment] = []
        trip_applied = False
        for segment in adjusted_segments:
            if not segment.end_time or private_trip["start_time"] >= segment.end_time or private_trip["end_time"] <= segment.start_time:
                next_segments.append(segment)
                continue

            trip_applied = True
            last_completion = find_last_completion_before(webtrack_summary, private_trip["start_time"], segment.start_time)
            split_end = last_completion["timestamp"] if last_completion else private_trip["start_time"]
            resume_time, resume_debug = infer_resume_after(private_trip["end_time"], gps_points, webtrack_summary)

            if split_end > segment.start_time:
                next_segments.append(
                    Segment(
                        start_time=segment.start_time,
                        end_time=split_end,
                        start_point=segment.start_point,
                        end_point=segment.end_point,
                        closure_reason="private_trip_cutoff_before_private_start",
                        is_estimated_end=False,
                        private_trip_split=True,
                    )
                )
            if resume_time and segment.end_time and resume_time < segment.end_time:
                resume_point = next(
                    (point for point in gps_points if point.timestamp >= resume_time),
                    None,
                )
                next_segments.append(
                    Segment(
                        start_time=resume_time,
                        end_time=segment.end_time,
                        start_point=_point_to_dict(resume_point) if resume_point else segment.start_point,
                        end_point=segment.end_point,
                        closure_reason="resumed_after_private_trip",
                        is_estimated_end=segment.is_estimated_end,
                        private_trip_split=True,
                    )
                )
            adjustments.append(
                {
                    "private_trip_id": private_trip["id"],
                    "reason": "segment_split_by_private_trip",
                    "segment_original_start": segment.start_time.isoformat(sep=" "),
                    "segment_original_end": segment.end_time.isoformat(sep=" ") if segment.end_time else None,
                    "adjusted_end_before_private": split_end.isoformat(sep=" "),
                    "resume_after_private": resume_time.isoformat(sep=" ") if resume_time else None,
                    "last_completion_before_private": last_completion["timestamp"].isoformat(sep=" ") if last_completion else None,
                    "resume_debug": resume_debug,
                }
            )

        adjusted_segments = sorted(next_segments, key=lambda item: item.start_time)
        if trip_applied:
            applied_private_trips.append(
                {
                    **private_trip,
                    "start_time": private_trip["start_time"].isoformat(sep=" "),
                    "end_time": private_trip["end_time"].isoformat(sep=" "),
                }
            )

    adjusted_segments = [segment for segment in adjusted_segments if segment.end_time and segment.end_time > segment.start_time]
    attach_segment_stop_ranges(adjusted_segments, webtrack_summary)
    return adjusted_segments, applied_private_trips, adjustments



def estimate_fallback_end(
    points: list[GPSPoint],
    webtrack_summary: dict[str, Any],
    home_latitude: float,
    home_longitude: float,
    last_order_latitude: float | None = None,
    last_order_longitude: float | None = None,
) -> dict[str, Any] | None:
    last_order = webtrack_summary.get("last_order_with_address")
    if last_order:
        base_time = last_order["timestamp"]
        if last_order_latitude is not None and last_order_longitude is not None:
            distance_home = haversine_meters(last_order_latitude, last_order_longitude, home_latitude, home_longitude)
            method = "last_order_address_plus_estimated_travel"
        else:
            nearest_point = min(points, key=lambda point: abs((point.timestamp - base_time).total_seconds()))
            distance_home = haversine_meters(nearest_point.latitude, nearest_point.longitude, home_latitude, home_longitude)
            method = "last_order_time_plus_estimated_travel_from_nearest_gps_point"
        travel_minutes = max(5, round((distance_home / 1000) / 35 * 60))
        estimated_end = base_time + timedelta(minutes=travel_minutes)
        return {
            "method": method,
            "timestamp": estimated_end,
            "estimated_travel_minutes": travel_minutes,
            "distance_home_m": round(distance_home, 1),
            "source_order": {
                "timestamp": base_time.isoformat(sep=" "),
                "raw_text": last_order["raw_text"],
                "address": last_order["extracted_address"],
            },
        }

    if points:
        return {
            "method": "last_gps_timestamp",
            "timestamp": points[-1].timestamp,
            "distance_home_m": round(points[-1].distance_to_home_m, 1),
            "source_order": None,
        }
    return None



def generate_movia_correction_text(result: dict[str, Any]) -> str:
    work_date = result["gps_summary"]["date"] or "Ikke fundet i filen"
    start_time = result["computed_start_time"] or "Ikke fundet i filen"
    end_time = result["computed_end_time"] or "Ikke fundet i filen"
    total_minutes = result["total_work_minutes"]
    segment_count = result["segment_count"]
    run_number = result["webtrack_summary"]["primary_run_number"] or "Ikke fundet i filen"
    order_number = result["webtrack_summary"]["primary_order_number"] or "Ikke fundet i filen"
    stop_numbers = result["webtrack_summary"]["stop_numbers"]
    stop_text = f"{stop_numbers[0]}-{stop_numbers[-1]} ({len(stop_numbers)} stop)" if stop_numbers else "Ikke fundet i filen"
    settlement = result.get("settlement_summary") or {}
    delay_summary = result.get("delay_summary") or {}

    lines = [
        f"Dato: {work_date}",
        f"Løbenummer: {run_number}",
        f"Ordre: {order_number}",
        f"Stoppesteder: {stop_text}",
        f"Starttid: {start_time}",
        f"Sluttid: {end_time}",
        f"Arbejdstid i alt: {total_minutes} minutter" if total_minutes is not None else "Arbejdstid i alt: Ikke oplyst",
        f"Antal forløb: {segment_count}",
        f"Sluttid er fastsat efter: {result['end_time_basis_label']}",
    ]
    if settlement:
        lines.extend(
            [
                f"Køretid i alt (by + land): {settlement.get('total_drive_minutes', 'Ikke oplyst')} minutter",
                f"Ventetid i alt (by + land): {settlement.get('total_wait_minutes', 'Ikke oplyst')} minutter",
                f"Afregnet min.: {settlement.get('afregnet_minutes', 'Ikke oplyst')}",
                f"Ønsket afregnet: {settlement.get('desired_minutes', 'Ikke oplyst')}",
                f"Difference: {settlement.get('difference_minutes', 'Ikke oplyst')} minutter",
            ]
        )
    if result.get("estimation_note"):
        lines.append(f"Bemærkning: {result['estimation_note']}")
    for private_trip in result.get("private_trips", []):
        lines.append(
            f"Privat kørsel fra {private_trip['start_time'][11:16]} til {private_trip['end_time'][11:16]} er fratrukket arbejdstiden."
        )

    main_delay = delay_summary.get("main_delay")
    if main_delay:
        lines.append(
            f"Forsinkelse vurderes primært ved stop {main_delay['stop_number']} ({main_delay['stop_type']}), ca. {main_delay['delay_only_minutes']} minutter. Årsag: {main_delay['reason']}."
        )

    lines.extend(
        [
            "Forklaring:",
            "Starttid er sat til det tidspunkt, hvor bilen forlader hjemmezonen efter stabil udkørsel.",
            "Sluttid er sat til første gyldige indkørsel i hjemmezonen og ikke til sidste GPS-punkt.",
            f"Systemet fandt {segment_count} forløb på dagen baseret på udkørsel, hjemkomst og eventuelle private ture.",
        ]
    )
    return "\n".join(lines)



def analyze_day(
    gps_path: str | Path,
    webtrack_path: str | Path,
    home_latitude: float | None = None,
    home_longitude: float | None = None,
    radius_meters: float = 300,
    min_return_dwell_minutes: int = 10,
    stable_point_count: int = 3,
    last_order_latitude: float | None = None,
    last_order_longitude: float | None = None,
    private_trip_overrides: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    gps_points = parse_gps_file(gps_path)
    webtrack_records = parse_webtrack_file(webtrack_path)
    webtrack_summary = summarize_webtrack(webtrack_records)

    home_source = "user_selected_coordinates"
    if home_latitude is None or home_longitude is None:
        home_latitude, home_longitude, home_source = derive_home_center_for_poc(gps_points)

    base_segments, segment_debug = detect_segments(
        gps_points,
        home_latitude=home_latitude,
        home_longitude=home_longitude,
        radius_meters=radius_meters,
        min_departure_points=stable_point_count,
        min_return_points=stable_point_count,
        min_return_dwell_minutes=min_return_dwell_minutes,
    )

    normalized_private_trips = normalize_private_trip_overrides(private_trip_overrides, gps_points[0].timestamp)
    attach_segment_stop_ranges(base_segments, webtrack_summary)
    adjusted_segments, applied_private_trips, segment_adjustments = apply_private_trip_overlays(
        base_segments,
        webtrack_summary,
        gps_points,
        normalized_private_trips,
    )

    final_segments = adjusted_segments or base_segments
    end_time_basis_label = "Første gyldige indkørsel i hjemmezone"
    estimation_note = None
    fallback = None
    needs_order_geocode = False

    if not final_segments:
        fallback = estimate_fallback_end(
            gps_points,
            webtrack_summary,
            home_latitude,
            home_longitude,
            last_order_latitude=last_order_latitude,
            last_order_longitude=last_order_longitude,
        )
        needs_order_geocode = bool(
            webtrack_summary.get("last_order_with_address")
            and last_order_latitude is None
            and last_order_longitude is None
        )
        if fallback:
            final_segments = [
                Segment(
                    start_time=gps_points[0].timestamp,
                    end_time=fallback["timestamp"],
                    start_point=_point_to_dict(gps_points[0]),
                    end_point=None,
                    closure_reason=fallback["method"],
                    is_estimated_end=True,
                    stops=[],
                )
            ]
            end_time_basis_label = "Estimeret sluttid"
            estimation_note = (
                "GPS viste ikke en sikker hjemkomst. Sluttiden er estimeret ud fra sidste meningsfulde ordre og rejsetid."
                if fallback["method"] != "last_gps_timestamp"
                else "GPS viste ikke en sikker hjemkomst. Sluttiden er sat til sidste GPS-tid som sidste fallback."
            )

    computed_start = final_segments[0].start_time if final_segments else None
    computed_end = final_segments[-1].end_time if final_segments and final_segments[-1].end_time else None
    total_work_minutes = sum(
        max(0, round((segment.end_time - segment.start_time).total_seconds() / 60))
        for segment in final_segments
        if segment.end_time
    )
    settlement_source = webtrack_summary.get("final_summary") or {}
    settlement_summary = {
        **settlement_source,
        "desired_minutes": total_work_minutes if final_segments else None,
        "difference_minutes": (total_work_minutes - settlement_source.get("afregnet_minutes"))
        if settlement_source.get("afregnet_minutes") is not None and final_segments
        else None,
        "summary_record_time": webtrack_summary.get("final_summary_record_time").isoformat(sep=" ")
        if webtrack_summary.get("final_summary_record_time")
        else None,
    }
    delay_summary = analyze_delay_points(webtrack_summary, gps_points)
    segment_debug["private_trip_overrides_count"] = len(normalized_private_trips)
    segment_debug["segment_adjustments"] = segment_adjustments

    result = {
        "home_center": {
            "latitude": round(home_latitude, 6),
            "longitude": round(home_longitude, 6),
            "radius_meters": radius_meters,
            "source": home_source,
            "stable_point_count": stable_point_count,
            "return_dwell_minutes": min_return_dwell_minutes,
        },
        "gps_summary": {
            "point_count": len(gps_points),
            "date": gps_points[0].timestamp.strftime("%d-%m-%Y") if gps_points else None,
            "first_timestamp": gps_points[0].timestamp.isoformat(sep=" ") if gps_points else None,
            "last_timestamp": gps_points[-1].timestamp.isoformat(sep=" ") if gps_points else None,
        },
        "webtrack_summary": {
            "primary_run_number": webtrack_summary["primary_run_number"],
            "primary_order_number": webtrack_summary["primary_order_number"],
            "stop_numbers": webtrack_summary["stop_numbers"],
            "stop_count": webtrack_summary["stop_count"],
            "v_lobe_slut_times": webtrack_summary["v_lobe_slut_times"],
            "final_summary": settlement_source,
            "last_meaningful_event_time": webtrack_summary["last_meaningful_event"]["timestamp"].isoformat(sep=" ") if webtrack_summary["last_meaningful_event"] else None,
            "last_meaningful_event_text": webtrack_summary["last_meaningful_event"]["raw_text"] if webtrack_summary["last_meaningful_event"] else None,
            "last_order_with_address": {
                "timestamp": webtrack_summary["last_order_with_address"]["timestamp"].isoformat(sep=" "),
                "address": webtrack_summary["last_order_with_address"]["extracted_address"],
                "raw_text": webtrack_summary["last_order_with_address"]["raw_text"],
            }
            if webtrack_summary["last_order_with_address"]
            else None,
        },
        "segment_count": len(final_segments),
        "segments": [
            {
                "start_time": segment.start_time.isoformat(sep=" "),
                "end_time": segment.end_time.isoformat(sep=" ") if segment.end_time else None,
                "start_point": segment.start_point,
                "end_point": segment.end_point,
                "closure_reason": segment.closure_reason,
                "is_estimated_end": segment.is_estimated_end,
                "stops": segment.stops or [],
                "private_trip_split": segment.private_trip_split,
            }
            for segment in final_segments
        ],
        "private_trips": applied_private_trips,
        "private_trip_suggestions": [],
        "settlement_summary": settlement_summary,
        "delay_summary": delay_summary,
        "computed_start_time": computed_start.strftime("%H:%M") if computed_start else None,
        "computed_end_time": computed_end.strftime("%H:%M") if computed_end else None,
        "total_work_minutes": total_work_minutes if final_segments else None,
        "end_time_basis_label": end_time_basis_label,
        "estimation_note": estimation_note,
        "needs_order_geocode": needs_order_geocode,
        "segment_debug": segment_debug,
        "fallback": {**fallback, "timestamp": fallback["timestamp"].isoformat(sep=" ")} if fallback else None,
        "gps_points_for_map": [_point_to_dict(point) for point in gps_points],
        "movia_correction_text": "",
    }
    result["movia_correction_text"] = generate_movia_correction_text(result)
    return result



def dump_json(data: dict[str, Any], path: str | Path) -> None:
    Path(path).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
